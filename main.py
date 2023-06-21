"""
    Thesis: Modelling Covid-19 dynamics in schools: the effect of air cleaning systems
    Name: Tika van Bennekum
    Student number: 13392425
    Supervisors: Valeria Krzhizhanovskaya and Daniel Bonn

    Description: 
        Code of the thesis: Modelling Covid-19 dynamics in schools: the effect of air cleaning systems
    
    Source used for basic SIR model implementation:
    https://scientific-python.readthedocs.io/en/latest/notebooks_rst/3_Ordinary_Differential_Equations/02_Examples/Epidemic_model_SIR.html
"""

import numpy as np
import scipy.integrate as integrate
import matplotlib.pyplot as plt
import openpyxl
from matplotlib import gridspec
from SALib.sample import saltelli
from SALib.analyze import sobol
from SALib.test_functions import Ishigami


def data():
    """
    Data from the experiments of the classrooms.
    returns the days and the corresponding data for the group with air filters and the group without air filters. The corresponding data has the number of new infections on a certain date.
    """
    wb = openpyxl.load_workbook("data.xlsx")
    ws = wb.active
    group = [ws.cell(row=i, column=2).value for i in range(2, 119)]
    day = [ws.cell(row=i, column=3).value for i in range(2, 119)]
    days = [i for i in range(1, 61)]
    data_with = [0 for _ in range(60)]
    data_without = [0 for _ in range(60)]

    for i in range(117):
        if group[i] == 1:
            for j in range(day[i] - 1, 60):
                data_with[j] += 1
        elif group[i] == 2:
            for j in range(day[i] - 1, 60):
                data_without[j] += 1

    return (days, data_with, data_without)


def country_data():
    """
    Data of the daily new infection cases in the Netherlands corresponding to the days of the experiment.
    """
    wb = openpyxl.load_workbook("country_data.xlsx")
    ws = wb.active
    group = [ws.cell(row=i, column=8).value for i in range(2, 119)]
    t = np.linspace(1, 60, 60)
    group = np.array(group[1:61])
    return group


def airfilter(beta_world, beta_classroom, gamma_world, initial):
    """
    Model to simulate the group with air filters. To use an SI model the gamma value should be set to zero. To have a standard SIR model, the initial variable should be set to one. If it is chosen to incoperate country statistics in the model, the function name should be changed to "derivative_country" to calculate "res".
    """
    N = 120.0
    I0, R0 = initial, 0
    S0 = N - I0 - R0
    gamma_classroom = gamma_world
    tmax = 240
    Nt = 240
    t = np.linspace(0, tmax, (Nt + 1))
    country_dat = country_data()

    # the odeint makes t values that best fit the situation resulting in t not being whole numbers anymore
    # in order to differentiate between when a kid is at school we take a certain range around the 4 multiples.
    def beta_t(t):
        """
        The beta value is dependent on wheter the individual is in the classroom or
        outside of the classroom. The values are fitted so that an individual resides in
        the classroom 25% of the time when it is a weekday.
        """
        if abs(t / 7 - round(t / 7)) < 0.1429:
            return beta_world
        elif abs(t / 4 - round(t / 4)) < 0.125:
            return beta_classroom
        else:
            return beta_world

    def gamma_t(t):
        """
        The gamma value is dependent on wheter the individual is in the classroom or
        outside of the classroom. The values are fitted so that an individual resides in
        the classroom 25% of the time when it is a weekday.
        """
        if abs(t / 4 - round(t / 4)) < 0.125:
            return gamma_classroom
        else:
            return gamma_world

    def derivative(X, t, N, beta_t, gamma_t):
        """
        Function to calculate the dynamics of the disease.
        """
        S, I, R = X
        dotS = -beta_t(t) * S * I / N
        dotI = beta_t(t) * S * I / N - gamma_t(t) * I
        dotR = gamma_t(t) * I
        return np.array([dotS, dotI, dotR])

    def derivative_country(X, t, N, beta_t, gamma_t):
        """
        Function to calculate the dynamics of the disease, incoperating the national infection statistics.
        """
        tc = int(np.rint(t))
        if tc > 59:
            tc = 59
        value = country_dat[tc] / 55822
        S, I, R = X
        dotS = -beta_t(t) * value * S * I / N
        dotI = beta_t(t) * value * S * I / N - gamma_t(t) * I
        dotR = gamma_t(t) * I
        return np.array([dotS, dotI, dotR])

    X0 = S0, I0, R0
    res = integrate.odeint(derivative, X0, t, args=(N, beta_t, gamma_t))
    S, I, R = res.T
    t = t // 4
    return (t, S, I, R, N)


def no_filter(beta_world, gamma_world, initial):
    """
    Model to simulate the group without air filters. To use an SI model the gamma value should be set to zero. To have a standard SIR model, the initial variable should be set to one. If it is chosen to incoperate country statistics in the model, the function name should be changed to "derivative_country" to calculate "res".
    """
    N = 240.0
    I0, R0 = initial, 0
    S0 = N - I0 - R0
    tmax = 60
    Nt = 60
    t = np.linspace(0, tmax, (Nt + 1))
    country_dat = country_data()

    def derivative(X, t):
        """
        Function to calculate the dynamics of the disease.
        """
        S, I, R = X
        dotS = -beta_world * S * I / N
        dotI = beta_world * S * I / N - gamma_world * I
        dotR = gamma_world * I
        return np.array([dotS, dotI, dotR])

    def derivative_country(X, t):
        """
        Function to calculate the dynamics of the disease, incoperating the national infection statistics.
        """
        tc = int(np.rint(t))
        if tc > 59:
            tc = 59
        value = country_dat[tc] / 55822
        S, I, R = X
        dotS = -beta_world * value * S * I / N
        dotI = beta_world * value * S * I / N - gamma_world * I
        dotR = gamma_world * I
        return np.array([dotS, dotI, dotR])

    X0 = S0, I0, R0
    res = integrate.odeint(derivative, X0, t)
    S, I, R = res.T
    return (t, S, I, R, N)


def rmse(actual, pred):
    """
    Function to calculate the root mean square error.
    """
    actual, pred = np.array(actual), np.array(pred)
    return np.sqrt(np.square(np.subtract(actual, pred)).mean())


def grid_search_filter():
    """
    Function to perform grid search for optimal model parameters for the model with air filter.
    """
    lowest_rmse = 1000.0
    beta_world = 100.0
    beta_classroom = 100.0
    gamma = 100.0
    initial = 0
    days, data_with, data_without = data()

    for i in np.arange(0.20, 0.70, 0.01):
        i = np.round(i, 2)
        print("i:", i)
        for j in np.arange(0.20, 0.70, 0.01):
            j = np.round(j, 2)
            for k in np.arange(1, 11, 0.5):
                k = np.round(k, 2)
                for n in range(1, 10):
                    n = float(n)
                    t_filter, S_filter, I_filter, R_filter, N_filter = airfilter(
                        i / 4, j, 1 / k, n
                    )

                    interpolated_infected_with_filter = np.interp(
                        np.linspace(1, 60, 60), t_filter, I_filter
                    )
                    interpolated_removed_with_filter = np.interp(
                        np.linspace(1, 60, 60), t_filter, R_filter
                    )
                    total_infections_interpolated = (
                        interpolated_infected_with_filter
                        + interpolated_removed_with_filter
                    )
                    rmse_value = rmse(data_with, total_infections_interpolated)

                    if rmse_value < lowest_rmse:
                        lowest_rmse = rmse_value
                        beta_world = i
                        beta_classroom = j
                        gamma = k
                        initial = n

    print(lowest_rmse)
    return (beta_world, beta_classroom, gamma, initial)


def grid_search_no_filter():
    """
    Function to perform grid search for optimal model parameters for the model without air filter.
    """
    difference_without_filter = 1000.0
    beta_world_without = 100.0
    gamma_world_without = 100.0
    initial = 0

    for i in np.arange(0.10, 0.70, 0.01):
        i = np.round(i, 2)
        print("i----", i)
        for j in np.arange(1, 12, 0.5):
            j = np.round(j, 2)
            for n in range(1, 11):
                n = float(n)
                (
                    t_no_filter,
                    S_no_filter,
                    I_no_filter,
                    R_no_filter,
                    N_no_filter,
                ) = no_filter(i, 1 / j, n)
                days, days_group1, days_group2 = data()

                interpolated_infected_without_filter = np.interp(
                    np.linspace(1, 60, 60), t_no_filter, I_no_filter
                )
                interpolated_removed_without_filter = np.interp(
                    np.linspace(1, 60, 60), t_no_filter, R_no_filter
                )
                total_infections = (
                    interpolated_infected_without_filter
                    + interpolated_removed_without_filter
                )
                new_difference_without = rmse(days_group2, total_infections)

                if new_difference_without < difference_without_filter:
                    difference_without_filter = new_difference_without
                    beta_world_without = i
                    gamma_world_without = j
                    initial = n

    print(difference_without_filter)
    return (beta_world_without, gamma_world_without, initial)


def sensitivity_with():
    """
    Sensitivity analysis for the model with air filter.
    Returns the first-order sobel indexes.
    """
    problem = {
        "num_vars": 4,
        "names": ["beta world", "beta class", "gamma", "initial"],
        "bounds": [[0.1, 0.70], [0.1, 0.70], [1, 10], [1, 10]],
    }

    param_values = saltelli.sample(problem, 1024)
    tmax = 60
    Nt = 60
    t = np.linspace(0, tmax, (Nt + 1))
    y = np.array([airfilter(*params) for params in param_values])
    sobol_indices = [sobol.analyze(problem, Y) for Y in y.T]
    S1s = np.array([s["S1"] for s in sobol_indices])

    return S1s


def sensitivity_without():
    """
    Sensitivity analysis for the model without air filter.
    Returns the first-order sobel indexes.
    """
    problem = {
        "num_vars": 3,
        "names": ["beta", "gamma", "initial"],
        "bounds": [
            [0.47, 0.50],
            [0.47, 0.50],
            [1, 3],
        ],
    }

    param_values = saltelli.sample(problem, 1024)
    tmax = 60
    Nt = 60
    t = np.linspace(0, tmax, (Nt + 1))
    y = np.array([no_filter(*params) for params in param_values])
    sobol_indices = [sobol.analyze(problem, Y) for Y in y.T]
    S1s = np.array([s["S1"] for s in sobol_indices])

    return S1s
